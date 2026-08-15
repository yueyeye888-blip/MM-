from __future__ import annotations

import copy
import json
import math
import shutil
import sqlite3
import threading
import time
import uuid
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


TZ = ZoneInfo("Asia/Singapore")


def now_dt() -> datetime:
    return datetime.now(TZ)


def iso_now() -> str:
    return now_dt().isoformat(timespec="milliseconds")


def number(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return float(value)
    try:
        result = float(value)
        return result if math.isfinite(result) else default
    except (TypeError, ValueError):
        return default


def normalize_name(value: Any) -> str:
    return " ".join(str(value or "").strip().split()).casefold()


def json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


@dataclass
class Settings:
    root: Path
    data: dict[str, Any]

    @classmethod
    def load(cls, path: Path) -> "Settings":
        data = json.loads(path.read_text(encoding="utf-8"))
        return cls(path.parent.resolve(), data)

    def path(self, key: str) -> Path:
        value = Path(self.data[key]).expanduser()
        return value.resolve() if value.is_absolute() else (self.root / value).resolve()

    def __getattr__(self, key: str) -> Any:
        try:
            return self.data[key]
        except KeyError as exc:
            raise AttributeError(key) from exc


class Database:
    def __init__(self, path: Path):
        path.parent.mkdir(parents=True, exist_ok=True)
        self.path = path
        self.lock = threading.RLock()
        self.conn = sqlite3.connect(path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        with self.lock:
            self.conn.executescript(
                """
                PRAGMA journal_mode=WAL;
                PRAGMA synchronous=NORMAL;
                CREATE TABLE IF NOT EXISTS raw_events (
                  event_id TEXT PRIMARY KEY, captured_at TEXT NOT NULL,
                  workbook_id TEXT, sheet_name TEXT, account_type TEXT,
                  account_name TEXT, field_key TEXT, cell_address TEXT,
                  old_value TEXT, new_value TEXT, change_type_cell TEXT,
                  bridge_session_id TEXT, source TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS stable_account_states (
                  state_id TEXT PRIMARY KEY, captured_at TEXT NOT NULL,
                  account_type TEXT NOT NULL, account_name TEXT NOT NULL,
                  normalized_name TEXT NOT NULL, payload TEXT NOT NULL,
                  is_valid INTEGER NOT NULL DEFAULT 1, correction_id TEXT,
                  quality_status TEXT NOT NULL, source_event_ids TEXT
                  ,source TEXT NOT NULL DEFAULT 'UNKNOWN'
                );
                CREATE INDEX IF NOT EXISTS idx_stable_account_time
                  ON stable_account_states(account_type, normalized_name, captured_at);
                CREATE TABLE IF NOT EXISTS global_states (
                  state_id TEXT PRIMARY KEY, captured_at TEXT NOT NULL,
                  payload TEXT NOT NULL, source TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS project_snapshots (
                  snapshot_id TEXT PRIMARY KEY, captured_at TEXT NOT NULL,
                  payload TEXT NOT NULL, quality_status TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_snapshot_time ON project_snapshots(captured_at);
                CREATE TABLE IF NOT EXISTS corrections (
                  correction_id TEXT PRIMARY KEY, captured_at TEXT NOT NULL,
                  account_type TEXT NOT NULL, account_name TEXT NOT NULL,
                  invalid_from TEXT, invalidated_count INTEGER NOT NULL,
                  details TEXT
                );
                CREATE TABLE IF NOT EXISTS backup_records (
                  backup_id TEXT PRIMARY KEY, captured_at TEXT NOT NULL,
                  source_mtime TEXT, backup_path TEXT, success INTEGER NOT NULL,
                  error TEXT
                );
                CREATE TABLE IF NOT EXISTS monitor_health (
                  event_id TEXT PRIMARY KEY, started_at TEXT NOT NULL,
                  ended_at TEXT, event_type TEXT NOT NULL, details TEXT
                );
                CREATE TABLE IF NOT EXISTS chat_queries (
                  query_id TEXT PRIMARY KEY, captured_at TEXT NOT NULL,
                  question TEXT NOT NULL, intent TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS chat_interactions (
                  interaction_id TEXT PRIMARY KEY, captured_at TEXT NOT NULL,
                  question TEXT NOT NULL, answer TEXT NOT NULL,
                  provider TEXT, model TEXT, template_id TEXT, segment_ids TEXT
                );
                """
            )
            self.conn.commit()

    def execute(self, sql: str, params: tuple = ()) -> sqlite3.Cursor:
        with self.lock:
            cur = self.conn.execute(sql, params)
            self.conn.commit()
            return cur

    def rows(self, sql: str, params: tuple = ()) -> list[sqlite3.Row]:
        with self.lock:
            return list(self.conn.execute(sql, params).fetchall())

    def row(self, sql: str, params: tuple = ()) -> sqlite3.Row | None:
        with self.lock:
            return self.conn.execute(sql, params).fetchone()


class WorkbookParser:
    """Header-driven, read-only parser for the shared market-making template."""

    SPOT_FIELDS = {
        "账户名称": "account_name", "状态": "status", "初始资金": "initial_capital",
        "累计追加": "added_capital", "现有资金": "current_funds", "现货数量(APR)": "position_qty",
        "变更类型": "change_type", "备注": "note", "账面成本金额": "excel_cost_basis",
        "成本均价": "excel_avg_cost", "监控字段ID": "monitor_id"
    }
    CONTRACT_FIELDS = {
        "账户名称": "account_name", "状态": "status", "初始资金": "initial_capital",
        "累计追加": "added_capital", "现有资金": "current_funds", "剩余可开": "available_funds",
        "方向": "direction", "持仓数量(APR)": "position_qty", "变更类型": "change_type", "备注": "note",
        "持仓均价": "excel_avg_entry", "已实现利润": "excel_realized", "未实现利润": "excel_unrealized"
    }

    def __init__(self, path: Path, sheet_name: str):
        self.path = path
        self.sheet_name = sheet_name

    @staticmethod
    def _labels(sheet, row: int) -> dict[str, int]:
        return {str(sheet.cell(row, c).value).strip(): c for c in range(1, sheet.max_column + 1)
                if sheet.cell(row, c).value not in (None, "")}

    @staticmethod
    def _find_header(sheet, required: str, start: int = 1) -> int:
        for row in range(start, sheet.max_row + 1):
            if str(sheet.cell(row, 1).value or "").strip() == required:
                return row
        raise ValueError(f"找不到表头: {required}")

    def _read_accounts(self, sheet, header_row: int, mapping: dict[str, str], account_type: str) -> list[dict]:
        labels = self._labels(sheet, header_row)
        columns = {dest: labels[src] for src, dest in mapping.items() if src in labels}
        if "position_qty" not in columns:
            prefix = "现货数量(" if account_type == "spot" else "持仓数量("
            dynamic = next((col for label, col in labels.items() if label.startswith(prefix)), None)
            if dynamic:
                columns["position_qty"] = dynamic
        if "account_name" not in columns or "position_qty" not in columns:
            raise ValueError(f"{account_type} 表头缺少账户名称或数量列")
        result = []
        for row in range(header_row + 1, sheet.max_row + 1):
            name = sheet.cell(row, columns["account_name"]).value
            first = str(name or "").strip()
            if first in ("现货合计", "合约合计"):
                break
            if not first:
                continue
            item = {key: sheet.cell(row, col).value for key, col in columns.items()}
            item.update({"account_type": account_type, "row": row})
            for key in ("initial_capital", "added_capital", "current_funds", "available_funds", "position_qty"):
                if key in item:
                    item[key] = number(item[key])
            item["status"] = str(item.get("status") or "启用").strip()
            item["change_type"] = str(item.get("change_type") or "正常").strip()
            item["account_name"] = first
            result.append(item)
        return result

    def parse(self) -> dict:
        # read_only=False is required for merged cells; this code never calls save().
        wb = load_workbook(self.path, data_only=True, read_only=False)
        sheet = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active
        global_labels = self._labels(sheet, 3)
        global_values = {label: sheet.cell(4, col).value for label, col in global_labels.items()}
        spot_header = self._find_header(sheet, "账户名称", 10)
        contract_header = self._find_header(sheet, "账户名称", spot_header + 2)
        return {
            "project": str(global_values.get("项目名称") or "APR"),
            "spot_price": number(global_values.get("现货市价")),
            "contract_price": number(global_values.get("合约市价")),
            "leverage": number(global_values.get("固定杠杆"), 2),
            # G7:H7 is merged in the shared template; openpyxl stores its value
            # on the top-left anchor G7 and returns None for H7.
            "sheet_spot_total_avg_cost": number(sheet["G7"].value),
            "sheet_spot_total_avg_cost_cell": f"{sheet.title}!G7:H7",
            "spots": self._read_accounts(sheet, spot_header, self.SPOT_FIELDS, "spot"),
            "contracts": self._read_accounts(sheet, contract_header, self.CONTRACT_FIELDS, "contract"),
            "workbook_name": self.path.name,
            "sheet_name": sheet.title,
            "source_mtime": self.path.stat().st_mtime,
        }

    def read_cells(self, addresses: list[str]) -> list[dict]:
        """Read saved cell values and resolve merged children to their anchor."""
        wb = load_workbook(self.path, data_only=True, read_only=False)
        sheet = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active
        result = []
        for raw_address in addresses:
            address = str(raw_address).strip().upper()
            resolved = address
            for merged in sheet.merged_cells.ranges:
                if address in merged:
                    resolved = sheet.cell(merged.min_row, merged.min_col).coordinate
                    break
            result.append({
                "sheet_name": sheet.title,
                "requested_address": address,
                "resolved_address": resolved,
                "value": sheet[resolved].value,
                "source": "SAVED_WORKBOOK",
            })
        return result


class MonitorEngine:
    GLOBAL_FIELDS = (
        "project", "spot_price", "contract_price", "leverage",
        "sheet_spot_total_avg_cost", "sheet_spot_total_avg_cost_cell",
    )
    RAW_FIELDS = {
        "spot": ("account_name", "status", "initial_capital", "added_capital", "current_funds", "position_qty", "change_type", "note"),
        "contract": ("account_name", "status", "initial_capital", "added_capital", "current_funds", "available_funds", "direction", "position_qty", "change_type", "note"),
    }

    def __init__(self, settings: Settings):
        self.settings = settings
        self.db = Database(settings.path("database_path"))
        self.parser = WorkbookParser(settings.path("source_workbook_path"), settings.sheet_name)
        self.lock = threading.RLock()
        self.raw: dict[str, Any] | None = None
        self.spots: dict[str, dict] = {}
        self.contracts: dict[str, dict] = {}
        self.global_state = {"project": "APR", "spot_price": 0.0, "contract_price": 0.0, "leverage": float(settings.leverage),
                             "sheet_spot_total_avg_cost": None, "sheet_spot_total_avg_cost_cell": None}
        self.pending: dict[str, tuple[float, dict, str, list[str]]] = {}
        self.started_at = iso_now()
        self.last_capture_at: str | None = None
        self.last_bridge_at: float | None = None
        self.last_bridge_session: str | None = None
        self.bridge_started_at: str | None = None
        self.bridge_last_snapshot_at: str | None = None
        self.bridge_last_event_at: str | None = None
        self.bridge_last_error: str | None = None
        self.bridge_event_registered = False
        self.bridge_workbook_name: str | None = None
        self.last_disk_mtime: float | None = None
        self.last_disk_signature: tuple[int, int, int] | None = None
        self.last_disk_poll_at: float | None = None
        self.last_disk_sync_at: str | None = None
        self.last_error: str | None = None
        self.has_unsaved_memory_changes = False
        self.correction_consumed: set[str] = set()
        self._bridge_was_online = False
        self._open_gap_id: str | None = None
        self.running = False
        self.thread: threading.Thread | None = None
        self._backup_hour: str | None = None

    def start(self) -> None:
        with self.lock:
            if self.running:
                return
            self.running = True
        self._bootstrap()
        self.thread = threading.Thread(target=self._loop, daemon=True, name="apr-monitor")
        self.thread.start()

    def stop(self) -> None:
        self.running = False

    def _bootstrap(self) -> None:
        try:
            snapshot = self.parser.parse()
            self.last_disk_mtime = snapshot.get("source_mtime")
            self.last_disk_signature = self._disk_signature()
            restored = self._restore_from_database()
            if restored:
                self.raw = {
                    **self.global_state,
                    "spots": [self._raw_from_state(x) for x in self.spots.values()],
                    "contracts": [self._raw_from_state(x) for x in self.contracts.values()],
                    "workbook_name": snapshot.get("workbook_name"), "sheet_name": snapshot.get("sheet_name")
                }
                self.observe(snapshot, "DISK_STARTUP_RECONCILE", immediate=True, record_unchanged=True)
            else:
                self.observe(snapshot, "DISK_BASELINE", immediate=True)
            self.last_error = None
        except Exception as exc:
            self.last_error = f"Excel 读取失败: {exc}"

    def _loop(self) -> None:
        next_poll = 0.0
        while self.running:
            try:
                now = time.time()
                if now >= next_poll:
                    self.poll_disk()
                    next_poll = now + max(1, int(self.settings.reconcile_interval_sec))
                self.flush_pending()
                self._track_bridge_gap()
                self.maybe_backup()
                self.last_error = None
            except Exception as exc:
                self.last_error = str(exc)
            time.sleep(0.25)

    def poll_disk(self) -> None:
        signature = self._disk_signature()
        self.last_disk_poll_at = time.time()
        if self.last_disk_signature is None or signature != self.last_disk_signature:
            snap = self.parser.parse()
            # Advance only after a successful read. A partially-written WPS save
            # will therefore be retried on the next one-second poll.
            self.last_disk_signature = signature
            self.last_disk_mtime = snap.get("source_mtime")
            self.observe(snap, "DISK_SAVE", immediate=True, record_unchanged=True)

    def _disk_signature(self) -> tuple[int, int, int]:
        stat = self.settings.path("source_workbook_path").stat()
        return stat.st_mtime_ns, stat.st_size, stat.st_ino

    def bridge_heartbeat(self, session_id: str | None, telemetry: dict | None = None) -> None:
        new_session = bool(session_id and session_id != self.last_bridge_session)
        if new_session:
            self.bridge_started_at = None
            self.bridge_last_snapshot_at = None
            self.bridge_last_event_at = None
            self.bridge_last_error = None
            self.bridge_event_registered = False
            self.bridge_workbook_name = None
        self.last_bridge_at = time.time()
        self.last_bridge_session = session_id or self.last_bridge_session
        if telemetry:
            self.bridge_started_at = telemetry.get("started_at")
            self.bridge_last_snapshot_at = telemetry.get("last_snapshot_at")
            self.bridge_last_event_at = telemetry.get("last_event_at")
            self.bridge_last_error = telemetry.get("last_error")
            self.bridge_event_registered = bool(telemetry.get("event_registered"))
            self.bridge_workbook_name = telemetry.get("workbook_name")

    def bridge_status(self) -> dict:
        health = self.health()
        return {
            "connected": health["status"] == "ONLINE",
            "mode": health["mode"],
            "session_id": self.last_bridge_session,
            "bridge_age_sec": health["bridge_age_sec"],
            "started_at": self.bridge_started_at,
            "last_snapshot_at": self.bridge_last_snapshot_at,
            "last_event_at": self.bridge_last_event_at,
            "last_error": self.bridge_last_error,
            "event_registered": self.bridge_event_registered,
            "workbook_name": self.bridge_workbook_name,
            "has_unsaved_memory_changes": self.has_unsaved_memory_changes,
        }

    def _restore_from_database(self) -> bool:
        rows = self.db.rows("SELECT payload FROM stable_account_states WHERE is_valid=1 ORDER BY captured_at DESC, rowid DESC")
        seen = set()
        for row in rows:
            state = json.loads(row["payload"])
            key = (state["account_type"], normalize_name(state["account_name"]))
            if key in seen:
                continue
            seen.add(key)
            target = self.spots if state["account_type"] == "spot" else self.contracts
            target[key[1]] = state
            if state.get("change_type") == "纠错":
                self.correction_consumed.add(f"{key[0]}:{key[1]}")
        global_row = self.db.row("SELECT payload FROM global_states ORDER BY captured_at DESC LIMIT 1")
        if global_row:
            self.global_state = json.loads(global_row["payload"])
        snapshot_row = self.db.row("SELECT captured_at FROM project_snapshots ORDER BY captured_at DESC LIMIT 1")
        if snapshot_row:
            self.last_capture_at = snapshot_row["captured_at"]
        return bool(seen or global_row)

    @staticmethod
    def _raw_from_state(state: dict) -> dict:
        keys = ("account_type", "account_name", "status", "initial_capital", "added_capital", "current_funds",
                "available_funds", "direction", "position_qty", "change_type", "note", "row")
        return {key: state.get(key) for key in keys if key in state}

    def _track_bridge_gap(self) -> None:
        online = self.last_bridge_at is not None and time.time() - self.last_bridge_at <= float(self.settings.offline_threshold_sec)
        if self._bridge_was_online and not online and self._open_gap_id is None:
            self._open_gap_id = str(uuid.uuid4())
            self.db.execute("INSERT INTO monitor_health VALUES (?,?,?,?,?)",
                            (self._open_gap_id, iso_now(), None, "MONITOR_GAP", "WPS Bridge heartbeat timeout"))
        elif not self._bridge_was_online and online and self._open_gap_id:
            self.db.execute("UPDATE monitor_health SET ended_at=? WHERE event_id=?", (iso_now(), self._open_gap_id))
            self._open_gap_id = None
        self._bridge_was_online = online

    def observe(
        self, snapshot: dict, source: str, immediate: bool = False,
        session_id: str | None = None, record_unchanged: bool = False,
    ) -> None:
        captured = iso_now()
        changed_any = False
        with self.lock:
            if source.startswith("DISK"):
                self.last_disk_sync_at = captured
            if source.startswith("WPS"):
                self.bridge_heartbeat(session_id)
                self.bridge_last_snapshot_at = captured
                self.bridge_workbook_name = snapshot.get("workbook_name") or self.bridge_workbook_name
                try:
                    disk = self.parser.parse()
                    self.has_unsaved_memory_changes = not self._same_inputs(snapshot, disk)
                except Exception:
                    self.has_unsaved_memory_changes = True
            if self.raw is None:
                self.raw = copy.deepcopy(snapshot)
                self.global_state = {k: snapshot.get(k) for k in self.GLOBAL_FIELDS}
                for item in snapshot.get("spots", []):
                    state = self._build_spot(item, None)
                    self._persist_state(state, source, captured, [])
                    self.spots[normalize_name(item["account_name"])] = state
                for item in snapshot.get("contracts", []):
                    state = self._build_contract(item, None)
                    self._persist_state(state, source, captured, [])
                    self.contracts[normalize_name(item["account_name"])] = state
                self._persist_global(source, captured)
                self.last_capture_at = captured
                self._persist_project_snapshot(captured)
                return

            changed_global = any(snapshot.get(k) != self.raw.get(k) for k in self.GLOBAL_FIELDS)
            if changed_global:
                changed_any = True
                ids = []
                for key in self.GLOBAL_FIELDS:
                    if snapshot.get(key) != self.raw.get(key):
                        ids.append(self._raw_event("global", "GLOBAL", key, self.raw.get(key), snapshot.get(key), source, captured, session_id))
                self.pending["global"] = (time.time() if immediate else time.time() + self.settings.debounce_ms / 1000, copy.deepcopy(snapshot), source, ids)

            for kind, list_key in (("spot", "spots"), ("contract", "contracts")):
                old_map = {normalize_name(x.get("account_name")): x for x in self.raw.get(list_key, [])}
                new_map = {normalize_name(x.get("account_name")): x for x in snapshot.get(list_key, [])}
                current_states = self.spots if kind == "spot" else self.contracts
                for key, item in new_map.items():
                    old = old_map.get(key, {})
                    changed = [f for f in self.RAW_FIELDS[kind] if item.get(f) != old.get(f)]
                    if changed:
                        changed_any = True
                        ids = [self._raw_event(kind, item["account_name"], f, old.get(f), item.get(f), source, captured, session_id) for f in changed]
                        self.pending[f"{kind}:{key}"] = (time.time() if immediate else time.time() + self.settings.debounce_ms / 1000, copy.deepcopy(item), source, ids)
                for key, old in old_map.items():
                    if key not in new_map and key in current_states and current_states[key].get("status") != "停用":
                        current_states[key].setdefault("quality_flags", []).append("STRUCTURE_ANOMALY")
            self.raw = copy.deepcopy(snapshot)
        if immediate:
            self.flush_pending()
            if record_unchanged and not changed_any:
                with self.lock:
                    self.last_capture_at = captured
                    self._persist_project_snapshot(captured)

    def _raw_event(self, kind: str, account: str, field: str, old: Any, new: Any, source: str, captured: str, session_id: str | None) -> str:
        event_id = str(uuid.uuid4())
        self.db.execute(
            "INSERT INTO raw_events VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (event_id, captured, self.raw.get("workbook_name") if self.raw else None,
             self.raw.get("sheet_name") if self.raw else None, kind, account, field, None,
             json_dumps(old), json_dumps(new), None, session_id, source),
        )
        return event_id

    def flush_pending(self) -> None:
        now = time.time()
        with self.lock:
            due = [(key, value) for key, value in self.pending.items() if value[0] <= now]
            for key, _ in due:
                del self.pending[key]
        if not due:
            return
        captured = iso_now()
        with self.lock:
            for key, (_, item, source, event_ids) in due:
                if key == "global":
                    self.global_state = {k: item.get(k) for k in self.GLOBAL_FIELDS}
                    self._persist_global(source, captured)
                    continue
                kind, norm = key.split(":", 1)
                states = self.spots if kind == "spot" else self.contracts
                prev = states.get(norm)
                correction_id = None
                correction_key = f"{kind}:{norm}"
                if item.get("change_type") != "纠错":
                    self.correction_consumed.discard(correction_key)
                elif correction_key not in self.correction_consumed:
                    correction_id = self._apply_correction(kind, item["account_name"], captured)
                    self.correction_consumed.add(correction_key)
                state = self._build_spot(item, prev) if kind == "spot" else self._build_contract(item, prev)
                state["correction_id"] = correction_id
                self._persist_state(state, source, captured, event_ids)
                states[norm] = state
            self.last_capture_at = captured
            self._persist_project_snapshot(captured)

    def _same_inputs(self, left: dict, right: dict) -> bool:
        for key in self.GLOBAL_FIELDS:
            if left.get(key) != right.get(key):
                return False
        for kind, list_key in (("spot", "spots"), ("contract", "contracts")):
            fields = self.RAW_FIELDS[kind]
            lm = {normalize_name(x.get("account_name")): x for x in left.get(list_key, [])}
            rm = {normalize_name(x.get("account_name")): x for x in right.get(list_key, [])}
            if set(lm) != set(rm):
                return False
            for name in lm:
                if any(lm[name].get(field) != rm[name].get(field) for field in fields):
                    return False
        return True

    def _base(self, item: dict) -> dict:
        initial = number(item.get("initial_capital"))
        added = number(item.get("added_capital"))
        return {
            "account_type": item["account_type"], "account_name": item["account_name"],
            "status": item.get("status") or "启用", "initial_capital": initial,
            "added_capital": added, "cumulative_capital": initial + added,
            "current_funds": number(item.get("current_funds")), "position_qty": number(item.get("position_qty")),
            "change_type": item.get("change_type") or "正常", "note": item.get("note"),
            "row": item.get("row"), "quality_flags": []
        }

    def _build_spot(self, item: dict, prev: dict | None) -> dict:
        state = self._base(item)
        qty = state["position_qty"]
        if not prev:
            cost = max(0.0, state["cumulative_capital"] - state["current_funds"])
            realized = 0.0
        else:
            cost = number(prev.get("spot_cost_basis"))
            realized = number(prev.get("spot_realized_pnl"))
            capital_delta = state["cumulative_capital"] - number(prev.get("cumulative_capital"))
            cash_delta = state["current_funds"] - number(prev.get("current_funds")) - capital_delta
            qty_delta = qty - number(prev.get("position_qty"))
            eps = 1e-9
            if qty_delta > eps and cash_delta < -eps:
                cost += -cash_delta
            elif qty_delta < -eps and cash_delta > eps:
                sell_qty = min(-qty_delta, number(prev.get("position_qty")))
                prev_avg = number(prev.get("spot_avg_cost"))
                realized += cash_delta - prev_avg * sell_qty
                cost = max(0.0, cost - prev_avg * sell_qty)
            elif abs(qty_delta) <= eps and abs(cash_delta) <= eps:
                pass
            elif abs(qty_delta) > eps and abs(cash_delta) <= eps:
                state["quality_flags"].append("DATA_INCOMPLETE")
            else:
                state["quality_flags"].append("DATA_INCONSISTENT")
        avg = cost / qty if qty > 0 else None
        market_value = qty * number(self.global_state.get("spot_price"))
        state.update({"spot_cost_basis": cost, "spot_avg_cost": avg, "spot_realized_pnl": realized,
                      "market_value": market_value, "spot_unrealized_pnl": market_value - cost})
        self._validate_common(state)
        return state

    def _build_contract(self, item: dict, prev: dict | None) -> dict:
        state = self._base(item)
        state["available_funds"] = number(item.get("available_funds"))
        state["direction"] = str(item.get("direction") or "").strip()
        qty = state["position_qty"]
        leverage = number(self.global_state.get("leverage"), number(self.settings.leverage, 2))
        if qty <= 0:
            avg = None
        elif prev and qty < number(prev.get("position_qty")) and state["direction"] == prev.get("direction"):
            avg = prev.get("contract_avg_entry")
        else:
            avg = ((state["cumulative_capital"] - state["available_funds"]) * leverage / qty) if qty else None
        price = number(self.global_state.get("contract_price"))
        unrealized = 0.0
        if avg is not None and state["direction"] == "多":
            unrealized = (price - avg) * qty
        elif avg is not None and state["direction"] == "空":
            unrealized = (avg - price) * qty
        liquidation = None
        if avg is not None and leverage:
            liquidation = avg * (1 - 1 / leverage) if state["direction"] == "多" else avg * (1 + 1 / leverage) if state["direction"] == "空" else None
        state.update({"contract_avg_entry": avg, "contract_realized_pnl": state["current_funds"] - state["cumulative_capital"],
                      "contract_unrealized_pnl": unrealized, "liquidation_estimate": liquidation})
        self._validate_common(state)
        if qty > 0 and state["direction"] not in ("多", "空"):
            state["quality_flags"].append("DATA_INCOMPLETE")
        return state

    @staticmethod
    def _validate_common(state: dict) -> None:
        if not state["account_name"]:
            state["quality_flags"].append("INVALID_ACCOUNT")
        for key in ("initial_capital", "added_capital", "current_funds", "position_qty"):
            if number(state.get(key)) < 0:
                state["quality_flags"].append("INVALID_NUMBER")

    def _persist_state(self, state: dict, source: str, captured: str, event_ids: list[str]) -> None:
        state["captured_at"] = captured
        quality = ",".join(sorted(set(state.get("quality_flags", [])))) or "OK"
        state["quality_status"] = quality
        self.db.execute(
            "INSERT INTO stable_account_states VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (str(uuid.uuid4()), captured, state["account_type"], state["account_name"], normalize_name(state["account_name"]),
             json_dumps(state), 1, state.get("correction_id"), quality, json_dumps(event_ids), source),
        )

    def _persist_global(self, source: str, captured: str) -> None:
        self.db.execute("INSERT INTO global_states VALUES (?,?,?,?)", (str(uuid.uuid4()), captured, json_dumps(self.global_state), source))

    def _apply_correction(self, kind: str, name: str, captured: str) -> str:
        correction_id = str(uuid.uuid4())
        rows = self.db.rows(
            "SELECT state_id,captured_at FROM stable_account_states WHERE account_type=? AND normalized_name=? AND is_valid=1 ORDER BY captured_at DESC",
            (kind, normalize_name(name)),
        )
        invalid = []
        gap = float(self.settings.correction_session_gap_sec)
        previous_time = now_dt()
        for row in rows:
            ts = datetime.fromisoformat(row["captured_at"])
            if (previous_time - ts).total_seconds() > gap and invalid:
                break
            invalid.append(row)
            previous_time = ts
        # Never erase the only bootstrap state; a correction needs a previous truth anchor.
        if len(invalid) > 1:
            invalid = invalid[:-1]
        else:
            invalid = []
        for row in invalid:
            self.db.execute("UPDATE stable_account_states SET is_valid=0, correction_id=? WHERE state_id=?", (correction_id, row["state_id"]))
        invalid_from = min((r["captured_at"] for r in invalid), default=None)
        self.db.execute("INSERT INTO corrections VALUES (?,?,?,?,?,?,?)",
                        (correction_id, captured, kind, name, invalid_from, len(invalid), "纠错标记已消费"))
        return correction_id

    def aggregate(self) -> dict:
        with self.lock:
            spots = [copy.deepcopy(x) for x in self.spots.values() if x.get("status") == "启用"]
            contracts = [copy.deepcopy(x) for x in self.contracts.values() if x.get("status") == "启用"]
            spot_qty = sum(x["position_qty"] for x in spots)
            spot_cost = sum(number(x.get("spot_cost_basis")) for x in spots)
            spot_realized = sum(number(x.get("spot_realized_pnl")) for x in spots)
            spot_market = spot_qty * number(self.global_state.get("spot_price"))
            longs = [x for x in contracts if x.get("direction") == "多"]
            shorts = [x for x in contracts if x.get("direction") == "空"]
            long_qty = sum(x["position_qty"] for x in longs)
            short_qty = sum(x["position_qty"] for x in shorts)
            long_weight = sum(x["position_qty"] * number(x.get("contract_avg_entry")) for x in longs)
            short_weight = sum(x["position_qty"] * number(x.get("contract_avg_entry")) for x in shorts)
            contract_realized = sum(number(x.get("contract_realized_pnl")) for x in contracts)
            contract_price = number(self.global_state.get("contract_price"))
            contract_unrealized = 0.0
            for account in contracts:
                avg = account.get("contract_avg_entry")
                if avg is None:
                    continue
                if account.get("direction") == "多":
                    contract_unrealized += (contract_price - number(avg)) * account["position_qty"]
                elif account.get("direction") == "空":
                    contract_unrealized += (number(avg) - contract_price) * account["position_qty"]
            net_qty = long_qty - short_qty
            project_net = spot_qty + net_qty
            spot_available_funds = sum(x["current_funds"] for x in spots)
            contract_available_funds = sum(x["current_funds"] for x in contracts)
            spot_cumulative_capital = sum(x["cumulative_capital"] for x in spots)
            spot_unrealized = spot_market - spot_cost
            # User-facing spot return is current account equity minus invested capital.
            # This is directly reconcilable to the Excel inputs and does not depend on
            # the internal historical cost-allocation process.
            spot_total_return = spot_available_funds + spot_market - spot_cumulative_capital
            contract_total_return = contract_realized + contract_unrealized
            realized = spot_realized + contract_realized
            unrealized = spot_unrealized + contract_unrealized
            total_return = spot_total_return + contract_total_return
            # Solve total project return at price P for zero:
            # spot_cash + P*spot_qty - spot_capital + contract_realized
            # + (P-long_avg)*long_qty + (short_avg-P)*short_qty = 0
            k_value = spot_cumulative_capital - spot_available_funds - contract_realized + long_weight - short_weight
            quality = sorted(set(flag for x in spots + contracts for flag in x.get("quality_flags", [])))
            if self._duplicates(spots) or self._duplicates(contracts):
                quality.append("DUPLICATE_ACCOUNT")
            return {
                "as_of": self.last_capture_at or self.started_at,
                "project_name": self.global_state.get("project") or "APR",
                "prices": {"spot": number(self.global_state.get("spot_price")), "contract": number(self.global_state.get("contract_price")), "leverage": number(self.global_state.get("leverage"), 2)},
                "capital": {"cumulative": sum(x["cumulative_capital"] for x in spots + contracts),
                            "available_funds": spot_available_funds + contract_available_funds,
                            "current_funds": spot_available_funds + contract_available_funds},
                "spot": {"qty": spot_qty, "cost_basis": spot_cost, "avg_cost": spot_cost / spot_qty if spot_qty else None,
                         "holding_total_avg_cost": self.global_state.get("sheet_spot_total_avg_cost"),
                         "holding_total_avg_cost_source": self.global_state.get("sheet_spot_total_avg_cost_cell") or "APR实时表!G7:H7",
                         "available_funds": spot_available_funds, "market_value": spot_market, "realized": spot_realized,
                         "unrealized": spot_unrealized, "total_return": spot_total_return, "accounts": spots},
                "contracts": {"long_qty": long_qty, "long_avg": long_weight / long_qty if long_qty else None,
                              "short_qty": short_qty, "short_avg": short_weight / short_qty if short_qty else None,
                              "gross_qty": long_qty + short_qty, "net_qty": net_qty,
                              "available_funds": contract_available_funds, "realized": contract_realized,
                              "unrealized": contract_unrealized, "total_return": contract_total_return, "accounts": contracts},
                "project": {"spot_total_return": spot_total_return, "contract_total_return": contract_total_return,
                            "realized": realized, "unrealized": unrealized, "total_return": total_return, "total_pnl": total_return,
                            "net_qty": project_net, "break_even": k_value / project_net if abs(project_net) > 1e-9 else None,
                            "break_even_status": "OK" if abs(project_net) > 1e-9 else "NO_SINGLE_BREAK_EVEN"},
                "data_quality": sorted(set(quality)),
            }

    @staticmethod
    def _duplicates(items: list[dict]) -> bool:
        names = [normalize_name(x["account_name"]) for x in items]
        return len(names) != len(set(names))

    def _persist_project_snapshot(self, captured: str) -> None:
        payload = self.aggregate()
        payload["as_of"] = captured
        quality = ",".join(payload["data_quality"]) or "OK"
        self.db.execute("INSERT INTO project_snapshots VALUES (?,?,?,?)", (str(uuid.uuid4()), captured, json_dumps(payload), quality))

    def health(self) -> dict:
        now = time.time()
        bridge_age = None if self.last_bridge_at is None else max(0.0, now - self.last_bridge_at)
        bridge_online = bridge_age is not None and bridge_age <= float(self.settings.offline_threshold_sec)
        poll_age = None if self.last_disk_poll_at is None else max(0.0, now - self.last_disk_poll_at)
        disk_online = poll_age is not None and poll_age <= max(3.0, float(self.settings.reconcile_interval_sec) * 3)
        status = "ONLINE" if disk_online and not self.last_error else "DEGRADED"
        lag = None
        if self.last_capture_at:
            lag = max(0, int((now_dt() - datetime.fromisoformat(self.last_capture_at)).total_seconds() * 1000))
        last_backup = self.db.row("SELECT * FROM backup_records ORDER BY captured_at DESC LIMIT 1")
        return {
            "status": status, "mode": "WPS_MEMORY" if bridge_online else "DISK_REALTIME",
            "as_of": self.last_capture_at, "lag_ms": lag, "bridge_age_sec": bridge_age,
            "last_disk_poll_at": datetime.fromtimestamp(self.last_disk_poll_at, TZ).isoformat() if self.last_disk_poll_at else None,
            "last_disk_sync_at": self.last_disk_sync_at,
            "database": "OK", "source_workbook": str(self.settings.path("source_workbook_path")),
            "source_mtime": datetime.fromtimestamp(self.last_disk_mtime, TZ).isoformat() if self.last_disk_mtime else None,
            "last_backup": dict(last_backup) if last_backup else None, "has_gap": not disk_online,
            "warnings": (["DISK_MONITOR_NOT_POLLING"] if not disk_online else []) +
                        (["UNSAVED_EXCEL"] if bridge_online and self.has_unsaved_memory_changes else []) +
                        ([self.last_error] if self.last_error else []),
        }

    def delta(self, window: str) -> dict:
        seconds = {"1m": 60, "5m": 300, "10m": 600}.get(window)
        if not seconds:
            raise ValueError("窗口必须是 1m、5m 或 10m")
        current = self.aggregate()
        current_at = datetime.fromisoformat(current["as_of"])
        target = current_at - timedelta(seconds=seconds)
        tolerance = int(self.settings.delta_tolerances_sec[window])
        rows = self.db.rows("SELECT captured_at,payload FROM project_snapshots WHERE captured_at BETWEEN ? AND ?",
                            ((target - timedelta(seconds=tolerance)).isoformat(), (target + timedelta(seconds=tolerance)).isoformat()))
        if not rows:
            return {"window": window, "status": "INSUFFICIENT_HISTORY", "as_of": current["as_of"], "reference_at": None, "offset_seconds": None, "delta": None}
        ref_row = min(rows, key=lambda r: abs((datetime.fromisoformat(r["captured_at"]) - target).total_seconds()))
        ref = json.loads(ref_row["payload"])
        gaps = self.db.rows("SELECT 1 FROM monitor_health WHERE event_type='DISK_MONITOR_GAP' AND started_at<=? AND (ended_at IS NULL OR ended_at>=?) LIMIT 1",
                            (current["as_of"], ref_row["captured_at"]))
        if self.health()["has_gap"] or gaps:
            return {"window": window, "status": "INCOMPLETE_GAP", "as_of": current["as_of"],
                    "reference_at": ref_row["captured_at"], "offset_seconds": None, "delta": None}
        metric_paths = {
            "cumulative_capital": ("capital", "cumulative"), "available_funds": ("capital", "available_funds"),
            "spot_qty": ("spot", "qty"), "spot_avg_cost": ("spot", "avg_cost"), "spot_realized": ("spot", "realized"),
            "spot_unrealized": ("spot", "unrealized"), "spot_total_return": ("spot", "total_return"),
            "long_qty": ("contracts", "long_qty"), "long_avg": ("contracts", "long_avg"),
            "short_qty": ("contracts", "short_qty"), "short_avg": ("contracts", "short_avg"), "contract_realized": ("contracts", "realized"),
            "contract_unrealized": ("contracts", "unrealized"), "contract_total_return": ("contracts", "total_return"),
            "total_return": ("project", "total_return"), "total_pnl": ("project", "total_pnl"), "break_even": ("project", "break_even")
        }
        delta = {}
        for label, path in metric_paths.items():
            a, b = current[path[0]].get(path[1]), ref[path[0]].get(path[1])
            delta[label] = None if a is None or b is None else number(a) - number(b)
        return {"window": window, "status": "OK", "as_of": current["as_of"], "reference_at": ref_row["captured_at"],
                "offset_seconds": (datetime.fromisoformat(ref_row["captured_at"]) - target).total_seconds(), "delta": delta}

    def maybe_backup(self, force: bool = False) -> dict | None:
        moment = now_dt()
        hour_key = moment.strftime("%Y-%m-%d-%H")
        if not force and (moment.minute != 0 or self._backup_hour == hour_key):
            return None
        source = self.settings.path("source_workbook_path")
        target_dir = self.settings.path("backup_dir")
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / f"APR_{moment.strftime('%Y-%m-%d_%H-%M')}.xlsx"
        ok, error = 1, None
        try:
            shutil.copy2(source, target)
        except Exception as exc:
            ok, error = 0, str(exc)
        captured = iso_now()
        self.db.execute("INSERT INTO backup_records VALUES (?,?,?,?,?,?)",
                        (str(uuid.uuid4()), captured, datetime.fromtimestamp(source.stat().st_mtime, TZ).isoformat(), str(target), ok, error))
        self._backup_hour = hour_key
        return {"success": bool(ok), "path": str(target) if ok else None, "error": error, "captured_at": captured}

    def history(self, start: str | None, end: str | None, limit: int = 200) -> list[dict]:
        sql, params = "SELECT captured_at,payload,quality_status FROM project_snapshots WHERE 1=1", []
        if start:
            sql += " AND captured_at>=?"; params.append(start)
        if end:
            sql += " AND captured_at<=?"; params.append(end)
        sql += " ORDER BY captured_at DESC LIMIT ?"; params.append(min(limit, 1000))
        return [{"captured_at": r["captured_at"], "quality_status": r["quality_status"], "summary": json.loads(r["payload"])} for r in self.db.rows(sql, tuple(params))]

    def chat(self, question: str) -> dict:
        q = question.strip().lower()
        if "10分钟" in q: intent = "delta_10m"
        elif "5分钟" in q: intent = "delta_5m"
        elif "1分钟" in q: intent = "delta_1m"
        elif "现货" in q: intent = "spot"
        elif "合约" in q or "多单" in q or "空单" in q: intent = "contracts"
        elif "资金" in q or "投入" in q: intent = "funds"
        elif "平衡" in q or "保本" in q: intent = "breakeven"
        else: intent = "summary"
        self.db.execute("INSERT INTO chat_queries VALUES (?,?,?,?)", (str(uuid.uuid4()), iso_now(), question, intent))
        summary, health = self.aggregate(), self.health()
        if intent.startswith("delta_"):
            window = intent.split("_")[1]
            data = self.delta(window)
            if data["status"] != "OK":
                body = f"{window.replace('m', '分钟')}历史数据不足，无法可靠计算净变化。"
            else:
                d = data["delta"]
                body = f"{window.replace('m', '分钟')}净变化：总可用资金 {d['available_funds']:,.2f} USDT，现货总收益 {d['spot_total_return']:,.2f} USDT，合约总收益 {d['contract_total_return']:,.2f} USDT，项目总收益 {d['total_return']:,.2f} USDT。"
        elif intent == "spot":
            s = summary["spot"]; body = f"现货共 {s['qty']:,.2f} APR，可用资金 {s['available_funds']:,.2f} USDT，现货持仓总均价 {(s['holding_total_avg_cost'] or 0):.6f}（{s['holding_total_avg_cost_source']}），市值 {s['market_value']:,.2f} USDT，现货总收益 {s['total_return']:,.2f} USDT。"
        elif intent == "contracts":
            c = summary["contracts"]; body = f"合约多单 {c['long_qty']:,.2f} APR，空单 {c['short_qty']:,.2f} APR，Gross {c['gross_qty']:,.2f}，Net {c['net_qty']:,.2f}，可用资金 {c['available_funds']:,.2f} USDT，合约总收益 {c['total_return']:,.2f} USDT。"
        elif intent == "funds":
            c = summary["capital"]; body = f"项目累计投入 {c['cumulative']:,.2f} USDT，总可用资金 {c['available_funds']:,.2f} USDT（启用的现货+合约账户“现有资金”合计）。"
        elif intent == "breakeven":
            p, s, c = summary["project"], summary["spot"], summary["contracts"]
            be = "净敞口接近 0，无单一盈亏平衡价" if p["break_even"] is None else f"{p['break_even']:.6f}"
            body = f"现货均价 {(s['avg_cost'] or 0):.6f}，多单均价 {(c['long_avg'] or 0):.6f}，空单均价 {(c['short_avg'] or 0):.6f}，项目综合盈亏平衡价：{be}。"
        else:
            p, c, s = summary["project"], summary["capital"], summary["spot"]
            body = f"项目累计投入 {c['cumulative']:,.2f} USDT，总可用资金 {c['available_funds']:,.2f} USDT；现货 {s['qty']:,.2f} APR；现货总收益 {p['spot_total_return']:,.2f} USDT，合约总收益 {p['contract_total_return']:,.2f} USDT，项目总收益 {p['total_return']:,.2f} USDT。"
        lag = (health["lag_ms"] or 0) / 1000
        tail = f"\n\n最近抓取时间：{summary['as_of']}\n距最近抓取：{lag:.1f} 秒\n监控状态：{health['status']}\n磁盘保存监控：{'存在监控缺口' if health['has_gap'] else '完整'}"
        return {"intent": intent, "answer": body + tail, "data": summary}
