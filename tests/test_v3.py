import json
import shutil
import tempfile
import time
import unittest
import uuid
from pathlib import Path

from openpyxl import load_workbook

from app.chat_templates import ChatTemplateStore
from app.core import Settings
from app.detection import DetectionService
from app.gpt_client import DataAssistant
from app.manager import ProjectManager


SOURCE = Path("/Users/xingxiu/Desktop/APR_做市实时表_V3_美化版.xlsx")


class MultiProjectTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        cfg = {
            "source_workbook_path": str(SOURCE), "sheet_name": "APR实时表",
            "backup_dir": "backups", "database_path": "data/apr.db", "host": "127.0.0.1", "port": 8765,
            "debounce_ms": 0, "reconcile_interval_sec": 60, "heartbeat_interval_sec": 5, "offline_threshold_sec": 15,
            "correction_session_gap_sec": 60, "delta_tolerances_sec": {"1m": 10, "5m": 30, "10m": 60},
            "leverage": 2, "backup_retention_days": None, "anomaly_thresholds": {}
        }
        config_path = self.root / "config.json"
        config_path.write_text(json.dumps(cfg), encoding="utf-8")
        self.manager = ProjectManager(self.root, Settings.load(config_path))
        self.manager.start()

    def tearDown(self):
        self.manager.stop()
        self.tmp.cleanup()

    def test_register_second_project_and_catalog(self):
        second = self.root / "Second.xlsx"
        shutil.copy2(SOURCE, second)
        spec = self.manager.register(str(second), "Second")
        self.assertEqual(spec["project_id"], "second")
        self.assertEqual(len(self.manager.list_projects()), 2)
        self.manager.update_catalog([{"name": "Second.xlsx", "full_name": str(second), "compatible": True}])
        self.assertTrue(self.manager.open_workbooks()[0]["registered"])
        database_path = spec["database_path"]
        removed = self.manager.remove(spec["project_id"])
        self.assertEqual(removed["database_preserved"], database_path)
        self.assertEqual(len(self.manager.list_projects()), 1)
        restored = self.manager.register(str(second), "重新命名也应恢复原项目")
        self.assertEqual(restored["database_path"], database_path)
        self.assertEqual(restored["project_id"], spec["project_id"])
        self.assertTrue(restored["enabled"])

    def test_detection_segments_and_combination(self):
        service = DetectionService(self.root / "data" / "control.db", self.manager)
        task = service.create_task("apr", "test")
        running = service.start_segment("apr", task["task_id"])
        stopped = service.stop_segment(running["segment_id"])
        self.assertEqual(stopped["status"], "STOPPED")
        self.assertFalse(stopped["has_gap"])
        combined = service.combine([stopped["segment_id"]])
        self.assertEqual(combined["segment_ids"], [stopped["segment_id"]])
        self.assertGreater(len(combined["metrics"]), 5)
        labels = {item["label"] for item in stopped["report"]["metrics"]}
        self.assertNotIn("累计投入", labels)
        self.assertNotIn("现货价格", labels)
        self.assertIn("现货数量变化", labels)
        self.assertEqual(stopped["report"]["target"]["project_id"], "apr")

    def test_saved_workbook_is_captured_immediately(self):
        second = self.root / "Saved.xlsx"
        shutil.copy2(SOURCE, second)
        spec = self.manager.register(str(second), "Saved")
        engine = self.manager.engine(spec["project_id"])
        previous_capture = engine.aggregate()["as_of"]
        workbook = load_workbook(second)
        sheet = workbook["APR实时表"]
        price_col = next(
            column for column in range(1, sheet.max_column + 1)
            if sheet.cell(3, column).value == "现货市价"
        )
        expected = float(sheet.cell(4, price_col).value or 0) + 0.001
        sheet.cell(4, price_col).value = expected
        time.sleep(0.003)
        workbook.save(second)
        engine.poll_disk()
        current = engine.aggregate()
        self.assertEqual(current["prices"]["spot"], expected)
        self.assertGreater(current["as_of"], previous_capture)
        self.assertEqual(engine.health()["mode"], "DISK_REALTIME")

    def test_spot_stage_cost_is_derived_from_state_changes(self):
        service = DetectionService(self.root / "data" / "cost-control.db", self.manager)
        engine = self.manager.engine("apr")
        baseline_account = {"account_name": "CostTest", "position_qty": 100, "current_funds": 800,
                            "cumulative_capital": 1000, "spot_cost_basis": 200, "spot_avg_cost": 2}
        final_account = {"account_name": "CostTest", "position_qty": 150, "current_funds": 650,
                         "cumulative_capital": 1000, "spot_cost_basis": 350, "spot_avg_cost": 350 / 150,
                         "quality_status": "OK"}
        engine.db.execute(
            "INSERT INTO stable_account_states VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (str(uuid.uuid4()), "2026-08-16T10:00:01+08:00", "spot", "CostTest", "costtest",
             json.dumps(final_account), 1, None, "OK", "[]", "TEST"),
        )
        baseline = {"spot": {"qty": 100, "cost_basis": 200, "avg_cost": 2, "accounts": [baseline_account]}}
        final = {"spot": {"qty": 150, "cost_basis": 350, "avg_cost": 350 / 150,
                           "holding_total_avg_cost": 0.06, "accounts": [final_account]}}
        result = service._spot_cost_analysis(
            engine, baseline, final, "2026-08-16T10:00:00+08:00", "2026-08-16T10:00:02+08:00"
        )
        self.assertEqual(result["known_purchase_qty"], 50)
        self.assertEqual(result["known_purchase_cost"], 150)
        self.assertEqual(result["known_purchase_avg_cost"], 3)
        self.assertEqual(result["ending_position_qty"], 150)
        self.assertEqual(result["ending_holding_total_avg_cost"], 0.06)

    def test_chat_templates_can_be_customized(self):
        store = ChatTemplateStore(self.root / "data" / "chat_templates.json")
        self.assertGreaterEqual(len(store.list()), 3)
        created = store.save({
            "name": "我的汇报", "description": "测试", "data_source": "CURRENT",
            "instructions": "只输出总可用资金和项目总收益。", "include_accounts": False,
            "include_events": False, "max_output_tokens": 600,
        })
        updated = store.save({**created, "name": "我的汇报 V2"}, created["template_id"])
        self.assertEqual(updated["name"], "我的汇报 V2")
        store.delete(created["template_id"])
        self.assertFalse(any(x["template_id"] == created["template_id"] for x in store.list()))

    def test_spot_total_average_is_forced_to_saved_g7(self):
        service = DetectionService(self.root / "data" / "chat-control.db", self.manager)
        assistant = DataAssistant(self.manager, service)
        context = assistant._sheet_metric_context("apr", "请读取 G7 和 H7")
        expected = self.manager.engine("apr").parser.read_cells(["G7"])[0]["value"]
        shown = f"{float(expected):.12g}"
        self.assertEqual(context["spot_holding_total_avg_cost"], expected)
        self.assertEqual(
            assistant._enforce_sheet_metrics("现货持仓总均价：0.1363360086337764", "apr"),
            f"现货持仓总均价：{shown}",
        )
        mixed = "现货持仓总均价：0.9\n阶段买入现货持仓均价：0.018"
        self.assertEqual(
            assistant._enforce_sheet_metrics(mixed, "apr"),
            f"现货持仓总均价：{shown}\n阶段买入现货持仓均价：0.018",
        )
        self.assertEqual(
            assistant._enforce_stage_purchase_avg(
                "阶段买入现货持仓均价：0.06",
                {"data": {"spot_cost_analysis": {"known_purchase_avg_cost": 0.018}}},
            ),
            "阶段买入现货持仓均价：0.018",
        )
        self.assertEqual(
            assistant._enforce_sheet_metrics("现货持仓均价：0.1363360086337764", "apr"),
            f"现货持仓均价：{shown}",
        )
        direct = assistant.ask("现货持仓总均价是多少？读取 G7 和 H7", "apr")
        self.assertEqual(direct["provider"], "LOCAL_VERIFIED_CELL")
        self.assertIn(f"现货持仓总均价：{shown}", direct["answer"])

    def test_gpt_56_terra_reasoning_profiles(self):
        service = DetectionService(self.root / "data" / "model-control.db", self.manager)
        assistant = DataAssistant(self.manager, service)
        payloads = []
        assistant.keys.get = lambda: "test-key"

        def fake_call(payload):
            payloads.append(payload)
            return {"id": "resp-test", "output_text": "已完成"}

        assistant._call_api = fake_call
        result = assistant.ask("当前项目情况", "apr")
        self.assertEqual(result["model"], "gpt-5.6-terra")
        self.assertEqual(payloads[-1]["reasoning"], {"effort": "low"})

        template = {
            "template_id": "test-report", "name": "测试报告", "data_source": "CURRENT",
            "instructions": "输出项目总收益。", "include_accounts": False,
            "include_events": False, "max_output_tokens": 600,
        }
        result = assistant.ask("", "apr", template)
        self.assertEqual(result["model"], "gpt-5.6-terra")
        self.assertEqual(payloads[-1]["reasoning"], {"effort": "medium"})
        self.assertEqual(payloads[-1]["max_output_tokens"], 2200)


if __name__ == "__main__":
    unittest.main()
